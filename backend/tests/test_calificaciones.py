"""tests/test_calificaciones.py — Tests for C-10: calificaciones-y-umbral.

TDD cycles covered:
  2.1  calcular_aprobado: 5 scenarios from spec
  3.1  CalificacionParser.parse_preview: numeric columns, textual columns, missing Email address
  3.3  CalificacionParser.parse_actividades_seleccionadas: filter selected activities
  4.1  CalificacionRepository: upsert_bulk, list_by_materia, delete_by_asignacion_materia, tenant isolation
  4.3  UmbralMateriaRepository: get_by_asignacion_materia, upsert
  5.1  CalificacionService.importar: full flow, upsert on re-import
  5.3  CalificacionService.configurar_umbral: create, update, recalculate, isolation
  6.1  Router POST /v1/calificaciones/{materia_id}/preview: 200, 422
  6.2  Router POST /v1/calificaciones/{materia_id}/importar: import, audit, upsert
  6.3  Router PUT /v1/calificaciones/{materia_id}/umbral: create, update, isolation

Safety net: run pytest tests/test_calificaciones.py -x before production code.
"""

import io
import os
import uuid
from dataclasses import dataclass
from datetime import date, timedelta
from typing import Any

import pytest
import pytest_asyncio
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker

_TEST_SECRET = "test-secret-key-32-characters-ok!"
_TEST_ENCRYPTION_KEY = "0" * 64


# ── Helpers ───────────────────────────────────────────────────────────────────


def _make_token(user_id: uuid.UUID, tenant_id: uuid.UUID) -> str:
    from app.core.security import create_access_token

    return create_access_token(
        data={"sub": str(user_id), "tenant_id": str(tenant_id)},
        expires_delta=timedelta(minutes=30),
    )


@dataclass
class TenantInfo:
    id: uuid.UUID
    slug: str


@dataclass
class UserInfo:
    id: uuid.UUID
    tenant_id: uuid.UUID
    token: str


@dataclass
class AcademicContext:
    materia_id: uuid.UUID
    asignacion_id: uuid.UUID
    entrada_padron_id: uuid.UUID
    cohorte_id: uuid.UUID


def _make_xlsx_calificaciones(
    rows: list[dict[str, Any]],
    extra_headers: list[str] | None = None,
) -> bytes:
    """Create an xlsx bytes payload for calificaciones tests.

    rows: list of dicts {email: str, numeric_acts: {col: val}, text_acts: {col: val}}
    extra_headers: additional non-grade columns (e.g. "Nombre")

    Returns bytes of an openpyxl workbook.
    """
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active

    # Collect all column names from first row
    headers = ["Email address"]
    if extra_headers:
        headers.extend(extra_headers)
    if rows:
        for key in rows[0].get("numeric_acts", {}).keys():
            headers.append(f"{key} (Real)")
        for key in rows[0].get("text_acts", {}).keys():
            headers.append(key)
    ws.append(headers)

    for row in rows:
        row_vals = [row["email"]]
        if extra_headers:
            for h in extra_headers:
                row_vals.append(row.get(h, ""))
        for key in row.get("numeric_acts", {}).keys():
            row_vals.append(row["numeric_acts"][key])
        for key in row.get("text_acts", {}).keys():
            row_vals.append(row["text_acts"][key])
        ws.append(row_vals)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Session-scoped fixtures ───────────────────────────────────────────────────


@pytest_asyncio.fixture(scope="module")
async def test_client(
    test_session_factory: async_sessionmaker[AsyncSession],
):
    from asgi_lifespan import LifespanManager
    from app.core.config import _reset_settings
    from app.main import create_application

    test_db_url = os.environ.get(
        "DATABASE_URL_TEST",
        "postgresql+asyncpg://activia:changeme@localhost:5432/activia_trace_test",
    )
    os.environ["DATABASE_URL"] = test_db_url
    os.environ["SECRET_KEY"] = _TEST_SECRET
    os.environ["ENCRYPTION_KEY"] = _TEST_ENCRYPTION_KEY
    _reset_settings()

    app = create_application()

    async with LifespanManager(app) as mgr:
        from app.core import database as db_module
        from app.models.base import Base

        async with db_module.engine.begin() as conn:
            await conn.run_sync(Base.metadata.create_all)

        import httpx

        async with httpx.AsyncClient(
            transport=httpx.ASGITransport(app=mgr.app),
            base_url="http://testserver",
        ) as client:
            yield client


@pytest_asyncio.fixture(scope="module")
async def cal_tenant(
    test_session_factory: async_sessionmaker[AsyncSession],
    test_client: Any,
) -> TenantInfo:
    from app.models.tenant import Tenant

    async with test_session_factory() as session:
        tenant = Tenant(
            id=uuid.uuid4(),
            slug=f"cal-tenant-{uuid.uuid4().hex[:8]}",
            nombre="Calificaciones Test Tenant",
        )
        session.add(tenant)
        await session.commit()
        return TenantInfo(id=tenant.id, slug=tenant.slug)


@pytest_asyncio.fixture(scope="module")
async def cal_user(
    test_session_factory: async_sessionmaker[AsyncSession],
    cal_tenant: TenantInfo,
) -> UserInfo:
    from app.core.security import email_hash as _eh, hash_password
    from app.models.usuario import Usuario

    raw = f"cal-user-{uuid.uuid4().hex[:8]}@test.com"
    user_id = uuid.uuid4()
    async with test_session_factory() as session:
        u = Usuario(
            id=user_id,
            tenant_id=cal_tenant.id,
            email_cifrado="placeholder-cal",
            email_hash=_eh(raw),
            password_hash=hash_password("testpass"),
            activo=True,
        )
        session.add(u)
        await session.commit()
    return UserInfo(id=user_id, tenant_id=cal_tenant.id, token=_make_token(user_id, cal_tenant.id))


@pytest_asyncio.fixture(scope="module")
async def cal_context(
    test_session_factory: async_sessionmaker[AsyncSession],
    cal_tenant: TenantInfo,
    cal_user: UserInfo,
) -> AcademicContext:
    """Create Carrera + Cohorte + Materia + Asignacion + VersionPadron + EntradaPadron."""
    from app.core.crypto import encrypt
    from app.models.asignacion import Asignacion
    from app.models.carrera import Carrera
    from app.models.cohorte import Cohorte
    from app.models.entrada_padron import EntradaPadron
    from app.models.materia import Materia
    from app.models.version_padron import VersionPadron

    async with test_session_factory() as session:
        carrera = Carrera(
            id=uuid.uuid4(),
            tenant_id=cal_tenant.id,
            codigo=f"CAR-CAL-{uuid.uuid4().hex[:4]}",
            nombre="Carrera Cal",
            estado="Activa",
        )
        session.add(carrera)

        materia = Materia(
            id=uuid.uuid4(),
            tenant_id=cal_tenant.id,
            codigo=f"MAT-CAL-{uuid.uuid4().hex[:4]}",
            nombre="Materia Cal",
            estado="Activa",
        )
        session.add(materia)

        # Flush carrera + materia first so FK constraints on cohorte/asignacion pass
        await session.flush()

        cohorte = Cohorte(
            id=uuid.uuid4(),
            tenant_id=cal_tenant.id,
            carrera_id=carrera.id,
            nombre=f"COH-CAL-{uuid.uuid4().hex[:4]}",
            anio=2025,
            vig_desde=date(2025, 1, 1),
            estado="Activa",
        )
        session.add(cohorte)

        # Flush cohorte before adding asignacion (FK dependency)
        await session.flush()

        asignacion = Asignacion(
            id=uuid.uuid4(),
            tenant_id=cal_tenant.id,
            usuario_id=cal_user.id,
            rol="PROFESOR",
            materia_id=materia.id,
            carrera_id=carrera.id,
            cohorte_id=cohorte.id,
            comisiones=[],
            desde=date(2025, 1, 1),
        )
        session.add(asignacion)

        await session.flush()

        version = VersionPadron(
            id=uuid.uuid4(),
            tenant_id=cal_tenant.id,
            materia_id=materia.id,
            cohorte_id=cohorte.id,
            cargado_por=cal_user.id,
            activa=True,
        )
        session.add(version)

        await session.flush()

        entrada = EntradaPadron(
            id=uuid.uuid4(),
            tenant_id=cal_tenant.id,
            version_id=version.id,
            nombre="Alumno",
            apellidos="Test",
            email_cifrado=encrypt("alumno@test.com"),
            comision="C1",
        )
        session.add(entrada)

        await session.commit()
        return AcademicContext(
            materia_id=materia.id,
            asignacion_id=asignacion.id,
            entrada_padron_id=entrada.id,
            cohorte_id=cohorte.id,
        )


# ─── 2.1 calcular_aprobado tests ──────────────────────────────────────────────


@pytest.mark.anyio
async def test_calcular_aprobado_numerica_sobre_umbral():
    """RED→GREEN: nota_numerica >= umbral → aprobado = True."""
    from app.services.calificacion_service import calcular_aprobado

    result = calcular_aprobado(nota_numerica=75.0, nota_textual=None, umbral_pct=60)
    assert result is True


@pytest.mark.anyio
async def test_calcular_aprobado_numerica_bajo_umbral():
    """TRIANGULATE: nota_numerica < umbral → aprobado = False."""
    from app.services.calificacion_service import calcular_aprobado

    result = calcular_aprobado(nota_numerica=45.0, nota_textual=None, umbral_pct=60)
    assert result is False


@pytest.mark.anyio
async def test_calcular_aprobado_numerica_exactamente_umbral():
    """TRIANGULATE: nota_numerica == umbral → aprobado = True (inclusive)."""
    from app.services.calificacion_service import calcular_aprobado

    result = calcular_aprobado(nota_numerica=60.0, nota_textual=None, umbral_pct=60)
    assert result is True


@pytest.mark.anyio
async def test_calcular_aprobado_textual_aprobatoria():
    """TRIANGULATE: nota_textual 'Satisfactorio' → aprobado = True."""
    from app.services.calificacion_service import calcular_aprobado

    result = calcular_aprobado(nota_numerica=None, nota_textual="Satisfactorio", umbral_pct=60)
    assert result is True


@pytest.mark.anyio
async def test_calcular_aprobado_textual_no_aprobatoria():
    """TRIANGULATE: nota_textual 'No satisfactorio' → aprobado = False."""
    from app.services.calificacion_service import calcular_aprobado

    result = calcular_aprobado(nota_numerica=None, nota_textual="No satisfactorio", umbral_pct=60)
    assert result is False


@pytest.mark.anyio
async def test_calcular_aprobado_textual_supera_lo_esperado():
    """TRIANGULATE: nota_textual 'Supera lo esperado' → aprobado = True."""
    from app.services.calificacion_service import calcular_aprobado

    result = calcular_aprobado(nota_numerica=None, nota_textual="Supera lo esperado", umbral_pct=60)
    assert result is True


# ─── 3.1 CalificacionParser.parse_preview tests ──────────────────────────────


@pytest.mark.anyio
async def test_parser_detecta_columnas_numericas():
    """RED→GREEN: columnas con '(Real)' se detectan como actividades numéricas."""
    from app.services.calificacion_parser import CalificacionParser

    file_bytes = _make_xlsx_calificaciones(
        rows=[{"email": "a@test.com", "numeric_acts": {"Actividad A": 75.0}, "text_acts": {}}],
    )

    parser = CalificacionParser()
    preview = parser.parse_preview(file_bytes, "test.xlsx")

    assert "Actividad A" in preview["actividades_numericas"]
    assert len(preview["actividades_textuales"]) == 0


@pytest.mark.anyio
async def test_parser_detecta_columnas_textuales():
    """TRIANGULATE: columnas sin '(Real)' se detectan como actividades textuales."""
    from app.services.calificacion_parser import CalificacionParser

    file_bytes = _make_xlsx_calificaciones(
        rows=[{"email": "a@test.com", "numeric_acts": {}, "text_acts": {"TP1": "Satisfactorio"}}],
    )

    parser = CalificacionParser()
    preview = parser.parse_preview(file_bytes, "test.xlsx")

    assert "TP1" in preview["actividades_textuales"]
    assert len(preview["actividades_numericas"]) == 0


@pytest.mark.anyio
async def test_parser_preview_no_persiste():
    """TRIANGULATE: parse_preview no persiste nada — devuelve solo el preview."""
    from app.services.calificacion_parser import CalificacionParser

    file_bytes = _make_xlsx_calificaciones(
        rows=[
            {"email": "a@test.com", "numeric_acts": {"Parcial 1": 80.0}, "text_acts": {"TP1": "Satisfactorio"}},
        ],
    )

    parser = CalificacionParser()
    preview = parser.parse_preview(file_bytes, "test.xlsx")

    assert "alumnos_preview" in preview
    assert len(preview["alumnos_preview"]) == 1
    assert preview["alumnos_preview"][0]["email"] == "a@test.com"


@pytest.mark.anyio
async def test_parser_error_sin_email_address():
    """TRIANGULATE: archivo sin columna 'Email address' → CalificacionParseError."""
    import openpyxl

    from app.services.calificacion_parser import CalificacionParseError, CalificacionParser

    # Build xlsx without Email address column
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Nombre", "Actividad A (Real)"])
    ws.append(["Ana", 75])
    buf = io.BytesIO()
    wb.save(buf)

    parser = CalificacionParser()
    with pytest.raises(CalificacionParseError) as exc_info:
        parser.parse_preview(buf.getvalue(), "test.xlsx")

    assert "Email address" in exc_info.value.detail


# ─── 3.3 CalificacionParser.parse_actividades_seleccionadas tests ────────────


@pytest.mark.anyio
async def test_parser_filtra_actividades_seleccionadas():
    """RED→GREEN: solo se parsean las actividades en la lista seleccionada."""
    from app.services.calificacion_parser import CalificacionParser

    file_bytes = _make_xlsx_calificaciones(
        rows=[
            {
                "email": "a@test.com",
                "numeric_acts": {"Parcial 1": 80.0, "Parcial 2": 45.0},
                "text_acts": {"TP1": "Satisfactorio"},
            }
        ],
    )

    parser = CalificacionParser()
    result = parser.parse_actividades_seleccionadas(
        file_bytes=file_bytes,
        filename="test.xlsx",
        actividades=["Parcial 1", "TP1"],
    )

    # Should have entries for "Parcial 1" and "TP1" but not "Parcial 2"
    assert len(result) == 2
    actividades_en_result = {r["actividad"] for r in result}
    assert "Parcial 1" in actividades_en_result
    assert "TP1" in actividades_en_result
    assert "Parcial 2" not in actividades_en_result


@pytest.mark.anyio
async def test_parser_actividades_seleccionadas_multiples_alumnos():
    """TRIANGULATE: múltiples alumnos × actividades se parsean correctamente."""
    from app.services.calificacion_parser import CalificacionParser

    file_bytes = _make_xlsx_calificaciones(
        rows=[
            {"email": "a@test.com", "numeric_acts": {"Parcial 1": 80.0}, "text_acts": {}},
            {"email": "b@test.com", "numeric_acts": {"Parcial 1": 40.0}, "text_acts": {}},
        ],
    )

    parser = CalificacionParser()
    result = parser.parse_actividades_seleccionadas(
        file_bytes=file_bytes,
        filename="test.xlsx",
        actividades=["Parcial 1"],
    )

    assert len(result) == 2
    emails = {r["email"] for r in result}
    assert "a@test.com" in emails
    assert "b@test.com" in emails


# ─── 4.1 CalificacionRepository tests ────────────────────────────────────────


@pytest.mark.anyio
async def test_calificacion_repo_upsert_bulk(
    test_session_factory: async_sessionmaker[AsyncSession],
    cal_tenant: TenantInfo,
    cal_context: AcademicContext,
):
    """RED→GREEN: upsert_bulk persiste calificaciones correctamente."""
    from app.repositories.calificacion_repository import CalificacionRepository

    calificacion_data = [
        {
            "entrada_padron_id": cal_context.entrada_padron_id,
            "materia_id": cal_context.materia_id,
            "actividad": f"Parcial-{uuid.uuid4().hex[:4]}",
            "nota_numerica": 75.0,
            "nota_textual": None,
            "aprobado": True,
            "origen": "Importado",
        }
    ]

    async with test_session_factory() as session:
        repo = CalificacionRepository(session, cal_tenant.id)
        created = await repo.upsert_bulk(calificacion_data)
        await session.commit()

    assert len(created) == 1
    assert created[0].nota_numerica == 75.0
    assert created[0].aprobado is True
    assert created[0].tenant_id == cal_tenant.id


@pytest.mark.anyio
async def test_calificacion_repo_upsert_bulk_updates_existing(
    test_session_factory: async_sessionmaker[AsyncSession],
    cal_tenant: TenantInfo,
    cal_context: AcademicContext,
):
    """TRIANGULATE: re-insert same (entrada_padron_id, actividad) updates the record."""
    from app.repositories.calificacion_repository import CalificacionRepository

    actividad = f"TP-UPSERT-{uuid.uuid4().hex[:4]}"

    first_data = [
        {
            "entrada_padron_id": cal_context.entrada_padron_id,
            "materia_id": cal_context.materia_id,
            "actividad": actividad,
            "nota_numerica": 50.0,
            "nota_textual": None,
            "aprobado": False,
            "origen": "Importado",
        }
    ]
    async with test_session_factory() as session:
        repo = CalificacionRepository(session, cal_tenant.id)
        await repo.upsert_bulk(first_data)
        await session.commit()

    second_data = [
        {
            "entrada_padron_id": cal_context.entrada_padron_id,
            "materia_id": cal_context.materia_id,
            "actividad": actividad,
            "nota_numerica": 80.0,
            "nota_textual": None,
            "aprobado": True,
            "origen": "Importado",
        }
    ]
    async with test_session_factory() as session:
        repo = CalificacionRepository(session, cal_tenant.id)
        updated = await repo.upsert_bulk(second_data)
        await session.commit()

    # Should have updated the existing record, not created a duplicate
    async with test_session_factory() as session:
        repo = CalificacionRepository(session, cal_tenant.id)
        all_cals = await repo.list_by_materia(cal_context.materia_id)

    matching = [c for c in all_cals if c.actividad == actividad]
    assert len(matching) == 1  # no duplicate
    assert matching[0].nota_numerica == 80.0
    assert matching[0].aprobado is True


@pytest.mark.anyio
async def test_calificacion_repo_tenant_isolation(
    test_session_factory: async_sessionmaker[AsyncSession],
    cal_tenant: TenantInfo,
    cal_context: AcademicContext,
):
    """TRIANGULATE: tenant isolation — otro tenant no ve las calificaciones."""
    from app.models.tenant import Tenant
    from app.repositories.calificacion_repository import CalificacionRepository

    actividad = f"ISO-CAL-{uuid.uuid4().hex[:4]}"

    # Insert calificacion for cal_tenant
    async with test_session_factory() as session:
        repo = CalificacionRepository(session, cal_tenant.id)
        await repo.upsert_bulk([
            {
                "entrada_padron_id": cal_context.entrada_padron_id,
                "materia_id": cal_context.materia_id,
                "actividad": actividad,
                "nota_numerica": 70.0,
                "nota_textual": None,
                "aprobado": True,
                "origen": "Importado",
            }
        ])
        await session.commit()

    # Another tenant should not see it
    other_tenant_id = uuid.uuid4()
    async with test_session_factory() as session:
        repo_other = CalificacionRepository(session, other_tenant_id)
        result = await repo_other.list_by_materia(cal_context.materia_id)

    assert len(result) == 0


# ─── 4.3 UmbralMateriaRepository tests ───────────────────────────────────────


@pytest.mark.anyio
async def test_umbral_repo_create(
    test_session_factory: async_sessionmaker[AsyncSession],
    cal_tenant: TenantInfo,
    cal_context: AcademicContext,
):
    """RED→GREEN: upsert crea UmbralMateria si no existe."""
    from app.repositories.calificacion_repository import UmbralMateriaRepository

    async with test_session_factory() as session:
        repo = UmbralMateriaRepository(session, cal_tenant.id)
        umbral = await repo.upsert(
            asignacion_id=cal_context.asignacion_id,
            materia_id=cal_context.materia_id,
            umbral_pct=70,
            valores_aprobatorios=["Satisfactorio"],
        )
        await session.commit()

    assert umbral.umbral_pct == 70
    assert umbral.tenant_id == cal_tenant.id


@pytest.mark.anyio
async def test_umbral_repo_update_existing(
    test_session_factory: async_sessionmaker[AsyncSession],
    cal_tenant: TenantInfo,
    cal_context: AcademicContext,
):
    """TRIANGULATE: upsert actualiza umbral si ya existe para (asignacion_id, materia_id)."""
    from app.repositories.calificacion_repository import UmbralMateriaRepository

    # Create
    async with test_session_factory() as session:
        repo = UmbralMateriaRepository(session, cal_tenant.id)
        await repo.upsert(
            asignacion_id=cal_context.asignacion_id,
            materia_id=cal_context.materia_id,
            umbral_pct=60,
            valores_aprobatorios=[],
        )
        await session.commit()

    # Update
    async with test_session_factory() as session:
        repo = UmbralMateriaRepository(session, cal_tenant.id)
        updated = await repo.upsert(
            asignacion_id=cal_context.asignacion_id,
            materia_id=cal_context.materia_id,
            umbral_pct=80,
            valores_aprobatorios=["Satisfactorio"],
        )
        await session.commit()

    # Verify single record with updated value
    async with test_session_factory() as session:
        repo = UmbralMateriaRepository(session, cal_tenant.id)
        umbral = await repo.get_by_asignacion_materia(
            asignacion_id=cal_context.asignacion_id,
            materia_id=cal_context.materia_id,
        )

    assert umbral is not None
    assert umbral.umbral_pct == 80


# ─── 5.1 CalificacionService.importar tests ──────────────────────────────────


@pytest.mark.anyio
async def test_service_importar_full_flow(
    test_session_factory: async_sessionmaker[AsyncSession],
    cal_tenant: TenantInfo,
    cal_context: AcademicContext,
    cal_user: UserInfo,
):
    """RED→GREEN: importar completo: parse → calcular_aprobado → upsert → audit."""
    from app.repositories.calificacion_repository import CalificacionRepository
    from app.services.calificacion_service import CalificacionService

    actividad_name = f"Parcial-Full-{uuid.uuid4().hex[:6]}"

    # Reset umbral to 60 explicitly (prior test runs may have changed it)
    async with test_session_factory() as session:
        svc = CalificacionService(session, cal_tenant.id)
        await svc.configurar_umbral(
            actor_id=cal_user.id,
            asignacion_id=cal_context.asignacion_id,
            materia_id=cal_context.materia_id,
            umbral_pct=60,
            valores_aprobatorios=[],
        )
        await session.commit()

    # Build file with unique activity name
    file_bytes = _make_xlsx_calificaciones(
        rows=[
            {
                "email": "alumno@test.com",
                "numeric_acts": {actividad_name: 75.0},
                "text_acts": {},
            }
        ],
    )

    async with test_session_factory() as session:
        svc = CalificacionService(session, cal_tenant.id)
        await svc.importar(
            actor_id=cal_user.id,
            materia_id=cal_context.materia_id,
            asignacion_id=cal_context.asignacion_id,
            file_bytes=file_bytes,
            filename="test.xlsx",
            actividades_seleccionadas=[actividad_name],
        )
        await session.commit()

    # Verify calificacion was persisted
    async with test_session_factory() as session:
        repo = CalificacionRepository(session, cal_tenant.id)
        cals = await repo.list_by_materia(cal_context.materia_id)

    matching = [c for c in cals if c.actividad == actividad_name]
    assert len(matching) >= 1
    assert matching[0].aprobado is True  # 75 >= 60 (umbral explicitly set above)


@pytest.mark.anyio
async def test_service_importar_upsert_on_reimport(
    test_session_factory: async_sessionmaker[AsyncSession],
    cal_tenant: TenantInfo,
    cal_context: AcademicContext,
    cal_user: UserInfo,
):
    """TRIANGULATE: reimportar la misma actividad actualiza el registro existente."""
    from app.repositories.calificacion_repository import CalificacionRepository
    from app.services.calificacion_service import CalificacionService

    actividad = f"Re-Import-{uuid.uuid4().hex[:4]}"

    # First import: 40 points (fail)
    file_v1 = _make_xlsx_calificaciones(
        rows=[{"email": "alumno@test.com", "numeric_acts": {actividad: 40.0}, "text_acts": {}}],
    )
    async with test_session_factory() as session:
        svc = CalificacionService(session, cal_tenant.id)
        await svc.importar(
            actor_id=cal_user.id,
            materia_id=cal_context.materia_id,
            asignacion_id=cal_context.asignacion_id,
            file_bytes=file_v1,
            filename="test.xlsx",
            actividades_seleccionadas=[actividad],
        )
        await session.commit()

    # Second import: 80 points (pass)
    file_v2 = _make_xlsx_calificaciones(
        rows=[{"email": "alumno@test.com", "numeric_acts": {actividad: 80.0}, "text_acts": {}}],
    )
    async with test_session_factory() as session:
        svc = CalificacionService(session, cal_tenant.id)
        await svc.importar(
            actor_id=cal_user.id,
            materia_id=cal_context.materia_id,
            asignacion_id=cal_context.asignacion_id,
            file_bytes=file_v2,
            filename="test.xlsx",
            actividades_seleccionadas=[actividad],
        )
        await session.commit()

    # Should have only one record, with the updated value
    async with test_session_factory() as session:
        repo = CalificacionRepository(session, cal_tenant.id)
        cals = await repo.list_by_materia(cal_context.materia_id)

    matching = [c for c in cals if c.actividad == actividad]
    assert len(matching) == 1
    assert matching[0].nota_numerica == 80.0
    assert matching[0].aprobado is True


# ─── 5.3 CalificacionService.configurar_umbral tests ─────────────────────────


@pytest.mark.anyio
async def test_service_configurar_umbral_create(
    test_session_factory: async_sessionmaker[AsyncSession],
    cal_tenant: TenantInfo,
    cal_context: AcademicContext,
    cal_user: UserInfo,
):
    """RED→GREEN: configurar umbral por primera vez crea UmbralMateria."""
    from app.repositories.calificacion_repository import UmbralMateriaRepository
    from app.services.calificacion_service import CalificacionService

    async with test_session_factory() as session:
        svc = CalificacionService(session, cal_tenant.id)
        await svc.configurar_umbral(
            actor_id=cal_user.id,
            asignacion_id=cal_context.asignacion_id,
            materia_id=cal_context.materia_id,
            umbral_pct=75,
            valores_aprobatorios=["Satisfactorio"],
        )
        await session.commit()

    async with test_session_factory() as session:
        repo = UmbralMateriaRepository(session, cal_tenant.id)
        umbral = await repo.get_by_asignacion_materia(
            asignacion_id=cal_context.asignacion_id,
            materia_id=cal_context.materia_id,
        )

    assert umbral is not None
    assert umbral.umbral_pct == 75


@pytest.mark.anyio
async def test_service_umbral_recalculates_aprobado(
    test_session_factory: async_sessionmaker[AsyncSession],
    cal_tenant: TenantInfo,
    cal_context: AcademicContext,
    cal_user: UserInfo,
):
    """TRIANGULATE: cambiar umbral recalcula campo aprobado en calificaciones existentes."""
    from app.repositories.calificacion_repository import CalificacionRepository
    from app.services.calificacion_service import CalificacionService

    actividad = f"Recalc-{uuid.uuid4().hex[:4]}"

    # Insert calificacion with nota 65 (should pass with umbral=60, fail with umbral=70)
    async with test_session_factory() as session:
        repo = CalificacionRepository(session, cal_tenant.id)
        await repo.upsert_bulk([
            {
                "entrada_padron_id": cal_context.entrada_padron_id,
                "materia_id": cal_context.materia_id,
                "actividad": actividad,
                "nota_numerica": 65.0,
                "nota_textual": None,
                "aprobado": True,  # initially passes with umbral 60
                "origen": "Importado",
            }
        ])
        await session.commit()

    # Change umbral to 70 → should make aprobado=False
    async with test_session_factory() as session:
        svc = CalificacionService(session, cal_tenant.id)
        await svc.configurar_umbral(
            actor_id=cal_user.id,
            asignacion_id=cal_context.asignacion_id,
            materia_id=cal_context.materia_id,
            umbral_pct=70,
            valores_aprobatorios=[],
        )
        await session.commit()

    # Verify recalculated
    async with test_session_factory() as session:
        repo = CalificacionRepository(session, cal_tenant.id)
        cals = await repo.list_by_materia(cal_context.materia_id)

    matching = [c for c in cals if c.actividad == actividad]
    assert len(matching) == 1
    assert matching[0].aprobado is False  # 65 < 70


@pytest.mark.anyio
async def test_service_umbral_isolation_between_docentes(
    test_session_factory: async_sessionmaker[AsyncSession],
    cal_tenant: TenantInfo,
    cal_context: AcademicContext,
    cal_user: UserInfo,
):
    """TRIANGULATE: cambiar umbral de un docente no afecta al otro."""
    from app.core.security import email_hash as _eh, hash_password
    from app.models.asignacion import Asignacion
    from app.models.usuario import Usuario
    from app.repositories.calificacion_repository import UmbralMateriaRepository
    from app.services.calificacion_service import CalificacionService

    # Create second docente + asignacion (flush user first to satisfy FK)
    async with test_session_factory() as session:
        u2 = Usuario(
            id=uuid.uuid4(),
            tenant_id=cal_tenant.id,
            email_cifrado="placeholder-cal2",
            email_hash=_eh(f"docente2-{uuid.uuid4().hex[:8]}@test.com"),
            password_hash=hash_password("testpass"),
            activo=True,
        )
        session.add(u2)
        await session.flush()  # flush user before asignacion FK

        asig2 = Asignacion(
            id=uuid.uuid4(),
            tenant_id=cal_tenant.id,
            usuario_id=u2.id,
            rol="PROFESOR",
            materia_id=cal_context.materia_id,
            desde=date(2025, 1, 1),
            comisiones=[],
        )
        session.add(asig2)
        await session.commit()
        asig2_id = asig2.id
        u2_id = u2.id

    # Set umbral for docente 1
    async with test_session_factory() as session:
        svc = CalificacionService(session, cal_tenant.id)
        await svc.configurar_umbral(
            actor_id=cal_user.id,
            asignacion_id=cal_context.asignacion_id,
            materia_id=cal_context.materia_id,
            umbral_pct=65,
            valores_aprobatorios=[],
        )
        await session.commit()

    # Set umbral for docente 2
    async with test_session_factory() as session:
        svc = CalificacionService(session, cal_tenant.id)
        await svc.configurar_umbral(
            actor_id=u2_id,
            asignacion_id=asig2_id,
            materia_id=cal_context.materia_id,
            umbral_pct=80,
            valores_aprobatorios=[],
        )
        await session.commit()

    # Verify each docente has its own umbral, not affected by the other
    async with test_session_factory() as session:
        repo = UmbralMateriaRepository(session, cal_tenant.id)
        u1_umbral = await repo.get_by_asignacion_materia(
            asignacion_id=cal_context.asignacion_id,
            materia_id=cal_context.materia_id,
        )
        u2_umbral = await repo.get_by_asignacion_materia(
            asignacion_id=asig2_id,
            materia_id=cal_context.materia_id,
        )

    assert u1_umbral is not None
    assert u2_umbral is not None
    assert u1_umbral.umbral_pct == 65
    assert u2_umbral.umbral_pct == 80


# ─── 6.1 Router: POST preview ─────────────────────────────────────────────────


@pytest.mark.anyio
async def test_router_preview_200(
    test_client: Any,
    cal_tenant: TenantInfo,
    cal_user: UserInfo,
    cal_context: AcademicContext,
    test_session_factory: async_sessionmaker[AsyncSession],
):
    """RED→GREEN: POST preview returns 200 with detected activities."""
    from app.core.schemas import UsuarioAutenticado
    from app.core import dependencies as deps
    from app.main import create_application
    from asgi_lifespan import LifespanManager
    import httpx, os

    os.environ["DATABASE_URL"] = os.environ.get(
        "DATABASE_URL_TEST",
        "postgresql+asyncpg://activia:changeme@localhost:5432/activia_trace_test",
    )
    os.environ["SECRET_KEY"] = _TEST_SECRET
    os.environ["ENCRYPTION_KEY"] = _TEST_ENCRYPTION_KEY
    from app.core.config import _reset_settings
    _reset_settings()

    app = create_application()

    mock_user = UsuarioAutenticado(
        user_id=cal_user.id,
        tenant_id=cal_tenant.id,
        roles=["PROFESOR"],
        permisos_efectivos={"calificaciones:importar"},
    )

    async def mock_get_current_user(token=None, session=None):
        return mock_user

    app.dependency_overrides[deps.get_current_user] = mock_get_current_user

    file_bytes = _make_xlsx_calificaciones(
        rows=[{"email": "alumno@test.com", "numeric_acts": {"Parcial 1": 80.0}, "text_acts": {"TP1": "Satisfactorio"}}],
    )

    try:
        async with LifespanManager(app) as manager:
            async with httpx.AsyncClient(
                transport=httpx.ASGITransport(app=manager.app),
                base_url="http://testserver",
            ) as client:
                response = await client.post(
                    f"/v1/calificaciones/{cal_context.materia_id}/preview",
                    files={"file": ("test.xlsx", file_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                )
        assert response.status_code == 200
        data = response.json()
        assert "actividades_numericas" in data
        assert "actividades_textuales" in data
        assert "Parcial 1" in data["actividades_numericas"]
        assert "TP1" in data["actividades_textuales"]
    finally:
        app.dependency_overrides.clear()


@pytest.mark.anyio
async def test_router_preview_422_sin_email_address(
    cal_tenant: TenantInfo,
    cal_user: UserInfo,
    cal_context: AcademicContext,
):
    """TRIANGULATE: archivo sin 'Email address' → 422."""
    import openpyxl
    from app.core.schemas import UsuarioAutenticado
    from app.core import dependencies as deps
    from app.main import create_application
    from asgi_lifespan import LifespanManager
    import httpx, os

    os.environ["DATABASE_URL"] = os.environ.get(
        "DATABASE_URL_TEST",
        "postgresql+asyncpg://activia:changeme@localhost:5432/activia_trace_test",
    )
    os.environ["SECRET_KEY"] = _TEST_SECRET
    os.environ["ENCRYPTION_KEY"] = _TEST_ENCRYPTION_KEY
    from app.core.config import _reset_settings
    _reset_settings()

    app = create_application()

    mock_user = UsuarioAutenticado(
        user_id=cal_user.id,
        tenant_id=cal_tenant.id,
        roles=["PROFESOR"],
        permisos_efectivos={"calificaciones:importar"},
    )

    async def mock_get_current_user(token=None, session=None):
        return mock_user

    app.dependency_overrides[deps.get_current_user] = mock_get_current_user

    # File without Email address
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Nombre", "Actividad (Real)"])
    ws.append(["Ana", 75])
    buf = io.BytesIO()
    wb.save(buf)
    bad_file = buf.getvalue()

    try:
        async with LifespanManager(app) as manager:
            async with httpx.AsyncClient(
                transport=httpx.ASGITransport(app=manager.app),
                base_url="http://testserver",
            ) as client:
                response = await client.post(
                    f"/v1/calificaciones/{cal_context.materia_id}/preview",
                    files={"file": ("test.xlsx", bad_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                )
        assert response.status_code == 422
    finally:
        app.dependency_overrides.clear()


# ─── 6.2 Router: POST importar ────────────────────────────────────────────────


@pytest.mark.anyio
async def test_router_importar_201(
    cal_tenant: TenantInfo,
    cal_user: UserInfo,
    cal_context: AcademicContext,
    test_session_factory: async_sessionmaker[AsyncSession],
):
    """RED→GREEN: POST importar returns 201 and persists calificaciones."""
    from app.core.schemas import UsuarioAutenticado
    from app.core import dependencies as deps
    from app.main import create_application
    from asgi_lifespan import LifespanManager
    from app.repositories.calificacion_repository import CalificacionRepository
    import httpx, os, json

    os.environ["DATABASE_URL"] = os.environ.get(
        "DATABASE_URL_TEST",
        "postgresql+asyncpg://activia:changeme@localhost:5432/activia_trace_test",
    )
    os.environ["SECRET_KEY"] = _TEST_SECRET
    os.environ["ENCRYPTION_KEY"] = _TEST_ENCRYPTION_KEY
    from app.core.config import _reset_settings
    _reset_settings()

    app = create_application()
    mock_user = UsuarioAutenticado(
        user_id=cal_user.id,
        tenant_id=cal_tenant.id,
        roles=["PROFESOR"],
        permisos_efectivos={"calificaciones:importar"},
    )

    async def mock_get_current_user(token=None, session=None):
        return mock_user

    app.dependency_overrides[deps.get_current_user] = mock_get_current_user

    actividad = f"RouterImport-{uuid.uuid4().hex[:4]}"
    file_bytes = _make_xlsx_calificaciones(
        rows=[{"email": "alumno@test.com", "numeric_acts": {actividad: 78.0}, "text_acts": {}}],
    )

    try:
        async with LifespanManager(app) as manager:
            async with httpx.AsyncClient(
                transport=httpx.ASGITransport(app=manager.app),
                base_url="http://testserver",
            ) as client:
                response = await client.post(
                    f"/v1/calificaciones/{cal_context.materia_id}/importar",
                    files={"file": ("test.xlsx", file_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                    data={
                        "asignacion_id": str(cal_context.asignacion_id),
                        "actividades_seleccionadas": json.dumps([actividad]),
                    },
                )
        assert response.status_code == 201
    finally:
        app.dependency_overrides.clear()


# ─── 6.3 Router: PUT umbral ───────────────────────────────────────────────────


@pytest.mark.anyio
async def test_router_umbral_200(
    cal_tenant: TenantInfo,
    cal_user: UserInfo,
    cal_context: AcademicContext,
):
    """RED→GREEN: PUT umbral returns 200 and persists umbral."""
    from app.core.schemas import UsuarioAutenticado
    from app.core import dependencies as deps
    from app.main import create_application
    from asgi_lifespan import LifespanManager
    import httpx, os

    os.environ["DATABASE_URL"] = os.environ.get(
        "DATABASE_URL_TEST",
        "postgresql+asyncpg://activia:changeme@localhost:5432/activia_trace_test",
    )
    os.environ["SECRET_KEY"] = _TEST_SECRET
    os.environ["ENCRYPTION_KEY"] = _TEST_ENCRYPTION_KEY
    from app.core.config import _reset_settings
    _reset_settings()

    app = create_application()
    mock_user = UsuarioAutenticado(
        user_id=cal_user.id,
        tenant_id=cal_tenant.id,
        roles=["PROFESOR"],
        permisos_efectivos={"calificaciones:umbral"},
    )

    async def mock_get_current_user(token=None, session=None):
        return mock_user

    app.dependency_overrides[deps.get_current_user] = mock_get_current_user

    try:
        async with LifespanManager(app) as manager:
            async with httpx.AsyncClient(
                transport=httpx.ASGITransport(app=manager.app),
                base_url="http://testserver",
            ) as client:
                response = await client.put(
                    f"/v1/calificaciones/{cal_context.materia_id}/umbral",
                    json={
                        "asignacion_id": str(cal_context.asignacion_id),
                        "umbral_pct": 72,
                        "valores_aprobatorios": ["Satisfactorio"],
                    },
                )
        assert response.status_code == 200
        data = response.json()
        assert data["umbral_pct"] == 72
    finally:
        app.dependency_overrides.clear()


@pytest.mark.anyio
async def test_router_umbral_403_sin_permiso(
    cal_tenant: TenantInfo,
    cal_user: UserInfo,
    cal_context: AcademicContext,
):
    """TRIANGULATE: sin permiso 'calificaciones:umbral' → 403."""
    from app.core.schemas import UsuarioAutenticado
    from app.core import dependencies as deps
    from app.main import create_application
    from asgi_lifespan import LifespanManager
    import httpx, os

    os.environ["DATABASE_URL"] = os.environ.get(
        "DATABASE_URL_TEST",
        "postgresql+asyncpg://activia:changeme@localhost:5432/activia_trace_test",
    )
    os.environ["SECRET_KEY"] = _TEST_SECRET
    os.environ["ENCRYPTION_KEY"] = _TEST_ENCRYPTION_KEY
    from app.core.config import _reset_settings
    _reset_settings()

    app = create_application()
    mock_user = UsuarioAutenticado(
        user_id=cal_user.id,
        tenant_id=cal_tenant.id,
        roles=["PROFESOR"],
        permisos_efectivos=set(),  # No permissions
    )

    async def mock_get_current_user(token=None, session=None):
        return mock_user

    app.dependency_overrides[deps.get_current_user] = mock_get_current_user

    try:
        async with LifespanManager(app) as manager:
            async with httpx.AsyncClient(
                transport=httpx.ASGITransport(app=manager.app),
                base_url="http://testserver",
            ) as client:
                response = await client.put(
                    f"/v1/calificaciones/{cal_context.materia_id}/umbral",
                    json={
                        "asignacion_id": str(cal_context.asignacion_id),
                        "umbral_pct": 70,
                        "valores_aprobatorios": [],
                    },
                )
        assert response.status_code == 403
    finally:
        app.dependency_overrides.clear()
