"""tests/test_padron.py — Tests for C-09: padron-ingesta-moodle.

TDD cycles covered:
  8.1  Versioning: activar nueva versión desactiva la anterior
  8.2  Primera carga crea versión activa
  8.3  Parser xlsx: columnas válidas y faltantes
  8.4  Parser csv: columnas válidas y faltantes
  8.5  EntradaPadron sin usuario_id (alumno sin cuenta)
  8.6  Tenant isolation
  8.7  MoodleWSClient mock: respuesta exitosa, 5xx, retry, no retry 4xx
  8.8  Fallback 502 en sync on-demand
  8.9  Vaciado scope-isolated
  8.10 RBAC: 403 sin permiso cargar/vaciar

Safety net: run pytest -x before adding production code to each section.
"""

import io
import os
import uuid
from dataclasses import dataclass
from datetime import timedelta
from unittest.mock import AsyncMock, MagicMock, patch

import pytest
import pytest_asyncio
from httpx import AsyncClient
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker

from app.models.base import Base
from app.models.carrera import Carrera
from app.models.cohorte import Cohorte
from app.models.materia import Materia
from app.models.tenant import Tenant
from app.models.usuario import Usuario

_TEST_SECRET = "test-secret-key-32-characters-ok!"
_TEST_ENCRYPTION_KEY = "0" * 64


# ── Helpers ───────────────────────────────────────────────────────────────────


def _make_token(user_id: uuid.UUID, tenant_id: uuid.UUID) -> str:
    from app.core.security import create_access_token

    return create_access_token(
        data={"sub": str(user_id), "tenant_id": str(tenant_id)},
        expires_delta=timedelta(minutes=30),
    )


@dataclass
class TenantInfo:
    id: uuid.UUID
    slug: str


@dataclass
class UserInfo:
    id: uuid.UUID
    tenant_id: uuid.UUID
    token: str


@dataclass
class AcademicContext:
    materia_id: uuid.UUID
    cohorte_id: uuid.UUID


# ── Session-scoped fixtures ───────────────────────────────────────────────────


@pytest_asyncio.fixture(scope="module")
async def test_client(
    test_session_factory: async_sessionmaker[AsyncSession],
) -> AsyncClient:
    from asgi_lifespan import LifespanManager
    from app.core.config import _reset_settings
    from app.main import create_application

    test_db_url = os.environ.get(
        "DATABASE_URL_TEST",
        "postgresql+asyncpg://activia:changeme@localhost:5432/activia_trace_test",
    )
    os.environ["DATABASE_URL"] = test_db_url
    os.environ["SECRET_KEY"] = _TEST_SECRET
    os.environ["ENCRYPTION_KEY"] = _TEST_ENCRYPTION_KEY
    _reset_settings()

    app = create_application()

    async with LifespanManager(app) as mgr:
        from app.core import database as db_module

        async with db_module.engine.begin() as conn:
            await conn.run_sync(Base.metadata.create_all)

        async with AsyncClient(
            transport=__import__("httpx").ASGITransport(app=mgr.app),
            base_url="http://testserver",
        ) as client:
            yield client


@pytest_asyncio.fixture(scope="module")
async def pd_tenant(
    test_session_factory: async_sessionmaker[AsyncSession],
    test_client: AsyncClient,
) -> TenantInfo:
    async with test_session_factory() as session:
        tenant = Tenant(
            id=uuid.uuid4(),
            slug=f"pd-tenant-{uuid.uuid4().hex[:8]}",
            nombre="Padron Test Tenant",
        )
        session.add(tenant)
        await session.commit()
        return TenantInfo(id=tenant.id, slug=tenant.slug)


@pytest_asyncio.fixture(scope="module")
async def pd_user(
    test_session_factory: async_sessionmaker[AsyncSession],
    pd_tenant: TenantInfo,
) -> UserInfo:
    from app.core.security import email_hash as _eh, hash_password

    raw = f"pd-user-{uuid.uuid4().hex[:8]}@test.com"
    user_id = uuid.uuid4()
    async with test_session_factory() as session:
        u = Usuario(
            id=user_id,
            tenant_id=pd_tenant.id,
            email_cifrado="placeholder-pd",
            email_hash=_eh(raw),
            password_hash=hash_password("testpass"),
            activo=True,
        )
        session.add(u)
        await session.commit()
    return UserInfo(id=user_id, tenant_id=pd_tenant.id, token=_make_token(user_id, pd_tenant.id))


@pytest_asyncio.fixture(scope="module")
async def pd_context(
    test_session_factory: async_sessionmaker[AsyncSession],
    pd_tenant: TenantInfo,
) -> AcademicContext:
    """Create a Materia + Cohorte + Carrera for tests."""
    async with test_session_factory() as session:
        carrera = Carrera(
            id=uuid.uuid4(),
            tenant_id=pd_tenant.id,
            codigo=f"CAR-{uuid.uuid4().hex[:6]}",
            nombre="Carrera Test",
            estado="Activa",
        )
        session.add(carrera)

        materia = Materia(
            id=uuid.uuid4(),
            tenant_id=pd_tenant.id,
            codigo=f"MAT-{uuid.uuid4().hex[:6]}",
            nombre="Materia Test",
            estado="Activa",
        )
        session.add(materia)

        cohorte = Cohorte(
            id=uuid.uuid4(),
            tenant_id=pd_tenant.id,
            carrera_id=carrera.id,
            nombre=f"COH-{uuid.uuid4().hex[:6]}",
            anio=2025,
            vig_desde=__import__("datetime").date(2025, 1, 1),
            estado="Activa",
        )
        session.add(cohorte)
        await session.commit()
        return AcademicContext(materia_id=materia.id, cohorte_id=cohorte.id)


# ─── 8.3 Parser xlsx tests ────────────────────────────────────────────────────


@pytest.mark.anyio
async def test_parser_xlsx_valid_columns():
    """RED→GREEN: valid xlsx with all required columns returns DTOs."""
    import openpyxl

    from app.services.padron_parser import PadronParser

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["nombre", "apellidos", "email", "comision", "regional"])
    ws.append(["Ana", "García", "ana@test.com", "C1", "Norte"])
    ws.append(["Luis", "Pérez", "luis@test.com", "C2", "Sur"])

    buf = io.BytesIO()
    wb.save(buf)
    file_bytes = buf.getvalue()

    parser = PadronParser()
    entradas = parser.parse_xlsx(file_bytes)

    assert len(entradas) == 2
    assert entradas[0].nombre == "Ana"
    assert entradas[0].apellidos == "García"
    assert entradas[0].email == "ana@test.com"
    assert entradas[0].comision == "C1"
    assert entradas[0].regional == "Norte"


@pytest.mark.anyio
async def test_parser_xlsx_missing_columns():
    """TRIANGULATE: xlsx with missing columns raises PadronParseError with detail."""
    import openpyxl

    from app.services.padron_parser import PadronParseError, PadronParser

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["nombre", "email"])  # missing apellidos, comision, regional
    ws.append(["Ana", "ana@test.com"])

    buf = io.BytesIO()
    wb.save(buf)

    parser = PadronParser()
    with pytest.raises(PadronParseError) as exc_info:
        parser.parse_xlsx(buf.getvalue())

    assert "apellidos" in exc_info.value.missing_columns
    assert "comision" in exc_info.value.missing_columns
    assert "regional" in exc_info.value.missing_columns


@pytest.mark.anyio
async def test_parser_xlsx_case_insensitive_headers():
    """TRIANGULATE: header matching is case-insensitive."""
    import openpyxl

    from app.services.padron_parser import PadronParser

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["NOMBRE", "APELLIDOS", "EMAIL", "COMISION", "REGIONAL"])
    ws.append(["Pedro", "López", "pedro@test.com", "C3", "Centro"])

    buf = io.BytesIO()
    wb.save(buf)

    parser = PadronParser()
    entradas = parser.parse_xlsx(buf.getvalue())

    assert len(entradas) == 1
    assert entradas[0].nombre == "Pedro"


# ─── 8.4 Parser csv tests ─────────────────────────────────────────────────────


@pytest.mark.anyio
async def test_parser_csv_valid_columns():
    """RED→GREEN: valid csv with all required columns returns DTOs."""
    from app.services.padron_parser import PadronParser

    csv_content = "nombre,apellidos,email,comision,regional\nAna,García,ana@test.com,C1,Norte\nLuis,Pérez,luis@test.com,C2,Sur\n"
    parser = PadronParser()
    entradas = parser.parse_csv(csv_content.encode("utf-8"))

    assert len(entradas) == 2
    assert entradas[0].nombre == "Ana"
    assert entradas[1].email == "luis@test.com"


@pytest.mark.anyio
async def test_parser_csv_missing_columns():
    """TRIANGULATE: csv missing required columns raises PadronParseError."""
    from app.services.padron_parser import PadronParseError, PadronParser

    csv_content = "nombre,email\nAna,ana@test.com\n"
    parser = PadronParser()

    with pytest.raises(PadronParseError) as exc_info:
        parser.parse_csv(csv_content.encode("utf-8"))

    assert "apellidos" in exc_info.value.missing_columns


@pytest.mark.anyio
async def test_parser_csv_empty_comision_treated_as_none():
    """TRIANGULATE: empty comision/regional in csv becomes None."""
    from app.services.padron_parser import PadronParser

    csv_content = "nombre,apellidos,email,comision,regional\nAna,García,ana@test.com,,\n"
    parser = PadronParser()
    entradas = parser.parse_csv(csv_content.encode("utf-8"))

    assert len(entradas) == 1
    assert entradas[0].comision is None
    assert entradas[0].regional is None


# ─── 8.7 MoodleWSClient tests (mock HTTP) ────────────────────────────────────


@pytest.mark.anyio
async def test_moodle_ws_success():
    """RED→GREEN: successful response returns list of MoodleEnrolledUser."""
    from app.integrations.moodle_ws import MoodleWSClient

    mock_response = MagicMock()
    mock_response.status_code = 200
    mock_response.json.return_value = [
        {
            "id": 1,
            "username": "agarcia",
            "firstname": "Ana",
            "lastname": "García",
            "email": "ana@test.com",
            "fullname": "Ana García",
        }
    ]

    with patch("httpx.AsyncClient") as mock_client_cls:
        mock_client = AsyncMock()
        mock_client_cls.return_value.__aenter__ = AsyncMock(return_value=mock_client)
        mock_client_cls.return_value.__aexit__ = AsyncMock(return_value=None)
        mock_client.get = AsyncMock(return_value=mock_response)

        client = MoodleWSClient(moodle_url="https://moodle.example.com", token="abc123")
        users = await client.get_enrolled_users(course_id=42)

    assert len(users) == 1
    assert users[0].firstname == "Ana"
    assert users[0].email == "ana@test.com"


@pytest.mark.anyio
async def test_moodle_ws_5xx_raises_error():
    """TRIANGULATE: HTTP 5xx raises MoodleWSError after retries."""
    from app.integrations.moodle_ws import MoodleWSClient, MoodleWSError

    mock_response = MagicMock()
    mock_response.status_code = 503
    mock_response.text = "Service Unavailable"

    with patch("httpx.AsyncClient") as mock_client_cls:
        mock_client = AsyncMock()
        mock_client_cls.return_value.__aenter__ = AsyncMock(return_value=mock_client)
        mock_client_cls.return_value.__aexit__ = AsyncMock(return_value=None)
        mock_client.get = AsyncMock(return_value=mock_response)

        # Speed up test — patch sleep to avoid actual waits
        with patch("asyncio.sleep", new_callable=AsyncMock):
            client = MoodleWSClient(moodle_url="https://moodle.example.com", token="abc123")
            with pytest.raises(MoodleWSError) as exc_info:
                await client.get_enrolled_users(course_id=42)

    assert exc_info.value.status_code is None or exc_info.value.status_code == 503


@pytest.mark.anyio
async def test_moodle_ws_4xx_no_retry():
    """TRIANGULATE: HTTP 4xx raises MoodleWSError immediately, no retry."""
    import httpx

    from app.integrations.moodle_ws import MoodleWSClient, MoodleWSError

    mock_response = MagicMock()
    mock_response.status_code = 403
    mock_response.text = "Forbidden"

    call_count = 0

    async def mock_get(*args, **kwargs):
        nonlocal call_count
        call_count += 1
        return mock_response

    with patch("httpx.AsyncClient") as mock_client_cls:
        mock_client = AsyncMock()
        mock_client_cls.return_value.__aenter__ = AsyncMock(return_value=mock_client)
        mock_client_cls.return_value.__aexit__ = AsyncMock(return_value=None)
        mock_client.get = mock_get

        client = MoodleWSClient(moodle_url="https://moodle.example.com", token="bad-token")
        with pytest.raises(MoodleWSError) as exc_info:
            await client.get_enrolled_users(course_id=42)

    # Should be called exactly once — no retry on 4xx
    assert call_count == 1
    assert exc_info.value.status_code == 403


@pytest.mark.anyio
async def test_moodle_ws_network_error_retries():
    """TRIANGULATE: network error triggers retry up to 3 times."""
    import httpx

    from app.integrations.moodle_ws import MoodleWSClient, MoodleWSError, _MAX_RETRIES

    call_count = 0

    async def mock_get(*args, **kwargs):
        nonlocal call_count
        call_count += 1
        raise httpx.ConnectError("Connection refused")

    with patch("httpx.AsyncClient") as mock_client_cls:
        mock_client = AsyncMock()
        mock_client_cls.return_value.__aenter__ = AsyncMock(return_value=mock_client)
        mock_client_cls.return_value.__aexit__ = AsyncMock(return_value=None)
        mock_client.get = mock_get

        with patch("asyncio.sleep", new_callable=AsyncMock):
            client = MoodleWSClient(moodle_url="https://moodle.example.com", token="abc")
            with pytest.raises(MoodleWSError):
                await client.get_enrolled_users(course_id=1)

    assert call_count == _MAX_RETRIES


@pytest.mark.anyio
async def test_moodle_ws_health_check_success():
    """TRIANGULATE: health_check returns True when WS responds."""
    from app.integrations.moodle_ws import MoodleWSClient

    mock_response = MagicMock()
    mock_response.status_code = 200
    mock_response.json.return_value = {"sitename": "Test Moodle"}

    with patch("httpx.AsyncClient") as mock_client_cls:
        mock_client = AsyncMock()
        mock_client_cls.return_value.__aenter__ = AsyncMock(return_value=mock_client)
        mock_client_cls.return_value.__aexit__ = AsyncMock(return_value=None)
        mock_client.get = AsyncMock(return_value=mock_response)

        client = MoodleWSClient(moodle_url="https://moodle.example.com", token="abc123")
        result = await client.health_check()

    assert result is True


@pytest.mark.anyio
async def test_moodle_ws_health_check_failure():
    """TRIANGULATE: health_check returns False when WS is down."""
    import httpx

    from app.integrations.moodle_ws import MoodleWSClient

    with patch("httpx.AsyncClient") as mock_client_cls:
        mock_client = AsyncMock()
        mock_client_cls.return_value.__aenter__ = AsyncMock(return_value=mock_client)
        mock_client_cls.return_value.__aexit__ = AsyncMock(return_value=None)
        mock_client.get = AsyncMock(side_effect=httpx.ConnectError("down"))

        with patch("asyncio.sleep", new_callable=AsyncMock):
            client = MoodleWSClient(moodle_url="https://moodle.example.com", token="abc123")
            result = await client.health_check()

    assert result is False


# ─── 8.1 Versioning tests (via repository) ───────────────────────────────────


@pytest.mark.anyio
async def test_versioning_new_version_deactivates_previous(
    test_session_factory: async_sessionmaker[AsyncSession],
    pd_tenant: TenantInfo,
    pd_user: UserInfo,
    pd_context: AcademicContext,
):
    """RED→GREEN: activar nueva versión desactiva la anterior."""
    from app.repositories.padron_repository import PadronRepository

    # Create first version
    async with test_session_factory() as session:
        repo = PadronRepository(session, pd_tenant.id)
        v1 = await repo.crear_version(
            materia_id=pd_context.materia_id,
            cohorte_id=pd_context.cohorte_id,
            cargado_por=pd_user.id,
        )
        await session.commit()
        v1_id = v1.id

    # Create second version
    async with test_session_factory() as session:
        repo = PadronRepository(session, pd_tenant.id)
        v2 = await repo.crear_version(
            materia_id=pd_context.materia_id,
            cohorte_id=pd_context.cohorte_id,
            cargado_por=pd_user.id,
        )
        await session.commit()
        v2_id = v2.id

    # Verify: v2 is active, v1 is not
    async with test_session_factory() as session:
        from sqlalchemy import select
        from app.models.version_padron import VersionPadron

        r1 = await session.get(VersionPadron, v1_id)
        r2 = await session.get(VersionPadron, v2_id)

    assert r1 is not None
    assert r2 is not None
    assert r1.activa is False  # previous version deactivated
    assert r2.activa is True   # new version is active
    assert r1.deleted_at is None  # not deleted, preserved for history
    assert r2.deleted_at is None


# ─── 8.2 First version creates active ────────────────────────────────────────


@pytest.mark.anyio
async def test_first_version_created_active(
    test_session_factory: async_sessionmaker[AsyncSession],
    pd_tenant: TenantInfo,
    pd_user: UserInfo,
):
    """RED→GREEN: primera carga de padrón crea versión activa."""
    from app.repositories.padron_repository import PadronRepository
    from app.models.carrera import Carrera
    from app.models.cohorte import Cohorte
    from app.models.materia import Materia
    import datetime

    # Create fresh materia + cohorte for this test (to avoid interference)
    async with test_session_factory() as session:
        carrera = Carrera(
            id=uuid.uuid4(),
            tenant_id=pd_tenant.id,
            codigo=f"C-FV-{uuid.uuid4().hex[:4]}",
            nombre="Car FV",
            estado="Activa",
        )
        session.add(carrera)
        materia = Materia(
            id=uuid.uuid4(),
            tenant_id=pd_tenant.id,
            codigo=f"M-FV-{uuid.uuid4().hex[:4]}",
            nombre="Mat FV",
            estado="Activa",
        )
        session.add(materia)
        cohorte = Cohorte(
            id=uuid.uuid4(),
            tenant_id=pd_tenant.id,
            carrera_id=carrera.id,
            nombre=f"COH-FV-{uuid.uuid4().hex[:4]}",
            anio=2025,
            vig_desde=datetime.date(2025, 1, 1),
            estado="Activa",
        )
        session.add(cohorte)
        await session.commit()
        mat_id = materia.id
        coh_id = cohorte.id

    async with test_session_factory() as session:
        repo = PadronRepository(session, pd_tenant.id)
        version = await repo.crear_version(
            materia_id=mat_id,
            cohorte_id=coh_id,
            cargado_por=pd_user.id,
        )
        await session.commit()

    assert version.activa is True
    assert version.materia_id == mat_id
    assert version.cohorte_id == coh_id


# ─── 8.5 EntradaPadron without usuario_id ────────────────────────────────────


@pytest.mark.anyio
async def test_entrada_padron_without_usuario_id(
    test_session_factory: async_sessionmaker[AsyncSession],
    pd_tenant: TenantInfo,
    pd_user: UserInfo,
    pd_context: AcademicContext,
):
    """RED→GREEN: EntradaPadron con usuario_id=None se persiste correctamente."""
    from app.repositories.padron_repository import PadronRepository
    from app.core.crypto import encrypt

    async with test_session_factory() as session:
        repo = PadronRepository(session, pd_tenant.id)
        version = await repo.crear_version(
            materia_id=pd_context.materia_id,
            cohorte_id=pd_context.cohorte_id,
            cargado_por=pd_user.id,
        )
        entradas = await repo.bulk_insert_entradas(
            version.id,
            [
                {
                    "nombre": "Sin",
                    "apellidos": "Cuenta",
                    "email_cifrado": encrypt("nocuenta@test.com"),
                    "usuario_id": None,
                    "comision": None,
                    "regional": None,
                }
            ],
        )
        await session.commit()
        entrada_id = entradas[0].id

    async with test_session_factory() as session:
        from app.models.entrada_padron import EntradaPadron
        ep = await session.get(EntradaPadron, entrada_id)

    assert ep is not None
    assert ep.usuario_id is None
    assert ep.nombre == "Sin"
    assert ep.apellidos == "Cuenta"


# ─── 8.6 Tenant isolation ─────────────────────────────────────────────────────


@pytest.mark.anyio
async def test_tenant_isolation_entradas(
    test_session_factory: async_sessionmaker[AsyncSession],
    pd_tenant: TenantInfo,
    pd_user: UserInfo,
    pd_context: AcademicContext,
):
    """RED→GREEN: entradas de tenant distinto no se mezclan."""
    from app.repositories.padron_repository import PadronRepository
    from app.core.crypto import encrypt
    import datetime

    # Create a second tenant
    async with test_session_factory() as session:
        tenant2 = Tenant(
            id=uuid.uuid4(),
            slug=f"pd-iso-{uuid.uuid4().hex[:8]}",
            nombre="Padron Isolation Tenant",
        )
        session.add(tenant2)

        user2 = Usuario(
            id=uuid.uuid4(),
            tenant_id=tenant2.id,
            email_cifrado="placeholder-iso",
            email_hash="hash-iso",
            password_hash="hash",
            activo=True,
        )
        session.add(user2)

        carrera2 = Carrera(
            id=uuid.uuid4(),
            tenant_id=tenant2.id,
            codigo=f"C-ISO-{uuid.uuid4().hex[:4]}",
            nombre="Car ISO",
            estado="Activa",
        )
        session.add(carrera2)

        materia2 = Materia(
            id=uuid.uuid4(),
            tenant_id=tenant2.id,
            codigo=f"M-ISO-{uuid.uuid4().hex[:4]}",
            nombre="Mat ISO",
            estado="Activa",
        )
        session.add(materia2)

        cohorte2 = Cohorte(
            id=uuid.uuid4(),
            tenant_id=tenant2.id,
            carrera_id=carrera2.id,
            nombre=f"COH-ISO-{uuid.uuid4().hex[:4]}",
            anio=2025,
            vig_desde=datetime.date(2025, 1, 1),
            estado="Activa",
        )
        session.add(cohorte2)
        await session.commit()
        t2_id = tenant2.id
        u2_id = user2.id
        m2_id = materia2.id
        c2_id = cohorte2.id

    # Create versions in each tenant
    async with test_session_factory() as session:
        repo1 = PadronRepository(session, pd_tenant.id)
        v1 = await repo1.crear_version(
            materia_id=pd_context.materia_id,
            cohorte_id=pd_context.cohorte_id,
            cargado_por=pd_user.id,
        )
        await repo1.bulk_insert_entradas(
            v1.id,
            [{"nombre": "T1", "apellidos": "User", "email_cifrado": encrypt("t1@test.com"), "usuario_id": None}],
        )

        repo2 = PadronRepository(session, t2_id)
        v2 = await repo2.crear_version(
            materia_id=m2_id,
            cohorte_id=c2_id,
            cargado_por=u2_id,
        )
        await repo2.bulk_insert_entradas(
            v2.id,
            [{"nombre": "T2", "apellidos": "User", "email_cifrado": encrypt("t2@test.com"), "usuario_id": None}],
        )
        await session.commit()
        v1_id = v1.id
        v2_id = v2.id

    # T1 repo cannot see T2's entries
    async with test_session_factory() as session:
        repo1 = PadronRepository(session, pd_tenant.id)
        entradas_t1 = await repo1.get_entradas(v1_id)
        entradas_t2_via_t1 = await repo1.get_entradas(v2_id)  # should be empty

    assert len(entradas_t1) >= 1
    assert all(e.nombre == "T1" for e in entradas_t1)
    assert len(entradas_t2_via_t1) == 0  # isolation enforced


# ─── 8.9 Vaciado scope-isolated ──────────────────────────────────────────────


@pytest.mark.anyio
async def test_vaciar_scope_isolated(
    test_session_factory: async_sessionmaker[AsyncSession],
    pd_tenant: TenantInfo,
    pd_user: UserInfo,
):
    """RED→GREEN: vaciado solo afecta materia×tenant del llamador."""
    from app.repositories.padron_repository import PadronRepository
    from app.models.version_padron import VersionPadron
    from app.core.crypto import encrypt
    import datetime

    # Create 2 materias in the same tenant
    async with test_session_factory() as session:
        carrera = Carrera(
            id=uuid.uuid4(),
            tenant_id=pd_tenant.id,
            codigo=f"C-VA-{uuid.uuid4().hex[:4]}",
            nombre="Car VA",
            estado="Activa",
        )
        session.add(carrera)
        mat_a = Materia(
            id=uuid.uuid4(),
            tenant_id=pd_tenant.id,
            codigo=f"M-VA-{uuid.uuid4().hex[:4]}",
            nombre="Mat VA A",
            estado="Activa",
        )
        mat_b = Materia(
            id=uuid.uuid4(),
            tenant_id=pd_tenant.id,
            codigo=f"M-VB-{uuid.uuid4().hex[:4]}",
            nombre="Mat VA B",
            estado="Activa",
        )
        session.add(mat_a)
        session.add(mat_b)
        cohorte = Cohorte(
            id=uuid.uuid4(),
            tenant_id=pd_tenant.id,
            carrera_id=carrera.id,
            nombre=f"COH-VA-{uuid.uuid4().hex[:4]}",
            anio=2025,
            vig_desde=datetime.date(2025, 1, 1),
            estado="Activa",
        )
        session.add(cohorte)
        await session.commit()
        mat_a_id = mat_a.id
        mat_b_id = mat_b.id
        coh_id = cohorte.id

    # Create 1 version per materia
    async with test_session_factory() as session:
        repo = PadronRepository(session, pd_tenant.id)
        va = await repo.crear_version(mat_a_id, coh_id, pd_user.id)
        vb = await repo.crear_version(mat_b_id, coh_id, pd_user.id)
        await repo.bulk_insert_entradas(va.id, [{"nombre": "A", "apellidos": "A", "email_cifrado": encrypt("a@t.com"), "usuario_id": None}])
        await repo.bulk_insert_entradas(vb.id, [{"nombre": "B", "apellidos": "B", "email_cifrado": encrypt("b@t.com"), "usuario_id": None}])
        await session.commit()
        va_id = va.id
        vb_id = vb.id

    # Vaciar only materia A
    async with test_session_factory() as session:
        repo = PadronRepository(session, pd_tenant.id)
        deleted = await repo.soft_delete_by_materia(mat_a_id)
        await session.commit()

    assert deleted > 0  # some rows deleted

    # Verify: version A is soft-deleted, version B is intact
    async with test_session_factory() as session:
        from sqlalchemy import select
        vra = await session.get(VersionPadron, va_id)
        vrb = await session.get(VersionPadron, vb_id)

    assert vra is not None and vra.deleted_at is not None  # deleted
    assert vrb is not None and vrb.deleted_at is None       # not affected


# ─── 8.8 Router: 502 when Moodle is down (via HTTP endpoint) ─────────────────


@pytest.mark.anyio
async def test_sync_moodle_502_when_moodle_down(
    pd_user: UserInfo,
    pd_context: AcademicContext,
    test_session_factory,
):
    """RED→GREEN: sync on-demand returns 502 when Moodle WS is unavailable."""
    import os
    from httpx import ASGITransport, AsyncClient as HClient
    from asgi_lifespan import LifespanManager
    from app.core.config import _reset_settings
    from app.core.schemas import UsuarioAutenticado
    from app.core import dependencies as deps
    from app.integrations.moodle_ws import MoodleWSError
    from app.main import create_application

    test_db_url = os.environ.get(
        "DATABASE_URL_TEST",
        "postgresql+asyncpg://activia:changeme@localhost:5432/activia_trace_test",
    )
    os.environ["DATABASE_URL"] = test_db_url
    os.environ["SECRET_KEY"] = _TEST_SECRET
    os.environ["ENCRYPTION_KEY"] = _TEST_ENCRYPTION_KEY
    _reset_settings()

    app = create_application()

    # Override current_user with a user who has padron:cargar
    mock_user = UsuarioAutenticado(
        user_id=pd_user.id,
        tenant_id=pd_user.tenant_id,
        roles=["PROFESOR"],
        permisos_efectivos={"padron:cargar"},
    )

    async def mock_get_current_user(token=None, session=None):
        return mock_user

    app.dependency_overrides[deps.get_current_user] = mock_get_current_user

    try:
        with patch(
            "app.services.padron_service.PadronService.sync_desde_moodle",
            new_callable=AsyncMock,
            side_effect=MoodleWSError("Connection refused"),
        ):
            async with LifespanManager(app) as manager:
                async with HClient(
                    transport=ASGITransport(app=manager.app),
                    base_url="http://testserver",
                ) as client:
                    response = await client.post(
                        f"/v1/padron/sync-moodle/{pd_context.materia_id}",
                        json={
                            "cohorte_id": str(pd_context.cohorte_id),
                            "course_id": 42,
                            "moodle_url": "https://moodle.example.com",
                            "moodle_token": "bad-token",
                        },
                        headers={"Authorization": f"Bearer {pd_user.token}"},
                    )

        assert response.status_code == 502
    finally:
        app.dependency_overrides.clear()


# ─── 8.10 RBAC tests ─────────────────────────────────────────────────────────


@pytest.mark.anyio
async def test_preview_without_cargar_permission(
    test_client: AsyncClient,
    pd_user: UserInfo,
):
    """RED→GREEN: POST /preview returns 403 without padron:cargar."""
    # User without any permissions hits the endpoint
    # Create a user without padron:cargar in a fresh tenant
    response = await test_client.post(
        "/v1/padron/preview",
        files={"file": ("test.csv", b"nombre,apellidos,email,comision,regional\n", "text/csv")},
        headers={"Authorization": f"Bearer {pd_user.token}"},
    )
    # pd_user doesn't have padron:cargar (no permissions seeded for test tenant)
    assert response.status_code == 403


@pytest.mark.anyio
async def test_vaciar_without_permission(
    test_client: AsyncClient,
    pd_user: UserInfo,
    pd_context: AcademicContext,
):
    """TRIANGULATE: DELETE /materia returns 403 without padron:vaciar."""
    response = await test_client.delete(
        f"/v1/padron/materia/{pd_context.materia_id}",
        headers={"Authorization": f"Bearer {pd_user.token}"},
    )
    assert response.status_code == 403


@pytest.mark.anyio
async def test_list_versiones_without_permission(
    test_client: AsyncClient,
    pd_user: UserInfo,
    pd_context: AcademicContext,
):
    """TRIANGULATE: GET /materia returns 403 without padron:leer."""
    response = await test_client.get(
        f"/v1/padron/materia/{pd_context.materia_id}",
        headers={"Authorization": f"Bearer {pd_user.token}"},
    )
    assert response.status_code == 403


# ─── 8.1 TRIANGULATE: multiple activations ───────────────────────────────────


@pytest.mark.anyio
async def test_versioning_three_versions_only_last_active(
    test_session_factory: async_sessionmaker[AsyncSession],
    pd_tenant: TenantInfo,
    pd_user: UserInfo,
):
    """TRIANGULATE: with 3 versions, only the last is active."""
    from app.repositories.padron_repository import PadronRepository
    from app.models.version_padron import VersionPadron
    import datetime

    async with test_session_factory() as session:
        carrera = Carrera(
            id=uuid.uuid4(),
            tenant_id=pd_tenant.id,
            codigo=f"C-3V-{uuid.uuid4().hex[:4]}",
            nombre="Car 3V",
            estado="Activa",
        )
        session.add(carrera)
        mat = Materia(
            id=uuid.uuid4(),
            tenant_id=pd_tenant.id,
            codigo=f"M-3V-{uuid.uuid4().hex[:4]}",
            nombre="Mat 3V",
            estado="Activa",
        )
        session.add(mat)
        coh = Cohorte(
            id=uuid.uuid4(),
            tenant_id=pd_tenant.id,
            carrera_id=carrera.id,
            nombre=f"COH-3V-{uuid.uuid4().hex[:4]}",
            anio=2025,
            vig_desde=datetime.date(2025, 1, 1),
            estado="Activa",
        )
        session.add(coh)
        await session.commit()
        mat_id = mat.id
        coh_id = coh.id

    version_ids = []
    for _ in range(3):
        async with test_session_factory() as session:
            repo = PadronRepository(session, pd_tenant.id)
            v = await repo.crear_version(mat_id, coh_id, pd_user.id)
            await session.commit()
            version_ids.append(v.id)

    async with test_session_factory() as session:
        versions = []
        for vid in version_ids:
            v = await session.get(VersionPadron, vid)
            versions.append(v)

    # Only last version is active
    active_versions = [v for v in versions if v.activa]
    assert len(active_versions) == 1
    assert active_versions[0].id == version_ids[-1]
